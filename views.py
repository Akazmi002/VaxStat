from django.shortcuts import render
from django.http import HttpResponse
import re
import openpyxl
from django.core.files.storage import default_storage
from django.http import JsonResponse



# Create your views here.


def list_entries():
    """
    Returns a list of all names of encyclopedia entries.
    """
    _, filenames = default_storage.listdir("exceldata")
    return list(sorted(re.sub(r"\.md$", "", filename)
                for filename in filenames if filename.endswith(".md")))





def index(request):
    return HttpResponse("Hello")





def getCounty(request, county):
    theFile = openpyxl.load_workbook('exceldata/COVID-19 Vaccine Data by County (2).xlsx')
    #print(theFile.sheetnames)
    currentSheet = theFile['By County']
    #print(currentSheet['B4'].value)


    for row in range(1, currentSheet.max_row + 1):
        for column in "A":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            VaccineData ="{}{}".format('F',row)
            Population="{}{}".format('G',row)
            Percentage  = (currentSheet[VaccineData].value)
            if currentSheet[cell_name].value == county:
                #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                #print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                #print ("County Name " + county)
                #print("Number of Vaccinations " + currentSheet[VaccineData].value)
                
                return JsonResponse({"CountyName": county, "VaccinesTaken":currentSheet[VaccineData].value, "Population 16+" : currentSheet[Population].value})
